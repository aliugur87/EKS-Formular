import customtkinter as ctk
import pandas as pd
import json
import os
from datetime import datetime, timedelta
from tkinter import filedialog, messagebox
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, asdict, field
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests
import threading
import re
import sys
import tempfile
import base64
import template_data # Az önce oluşturduğumuz dosyayı import ediyoruz
import sys
import os

# PyInstaller için path düzeltmesi
if getattr(sys, 'frozen', False):
    # PyInstaller bundle içindeyiz
    os.environ['CUSTOMTKINTER_PATH'] = sys._MEIPASS
    import customtkinter
    customtkinter.set_appearance_mode("dark")
    customtkinter.set_default_color_theme("blue")

APP_VERSION = "v1.0.0"


def resource_path(relative_path):
    """ Geliştirme ve PyInstaller için kaynaklara mutlak yol alır """
    try:
        # PyInstaller geçici bir klasör oluşturur ve yolu _MEIPASS içinde saklar
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Dil sistemi
LANGUAGES = {
    "DE": {
        "app_title": "EKS Formular Ausfüller Pro",
        "customer": "Kunde",
        "period": "Zeitraum",
        "template": "Vorlage", 
        "load_bwa": "BWA Datei laden",
        "auto_mapping": "Automatische Zuordnung",
        "export_eks": "EKS Exportieren",
        "new_customer": "Neuer Kunde",
        "customer_code": "Kundennummer",
        "customer_name": "Kundenname",
        "from_date": "Von Datum",
        "to_date": "Bis Datum",
        "quick_select": "Schnellauswahl",
        "mapping_results": "Zuordnungsergebnisse",
        "confidence": "Vertrauen",
        "monthly_values": "Monatswerte",
        "total": "Gesamt",
        "success": "Erfolgreich",
        "error": "Fehler",
        "settings": "Einstellungen",
        "api_key": "API Schlüssel",
        "save": "Speichern",
        "cancel": "Abbrechen",
        "loading": "Laden...",
        "file_loaded": "Datei geladen",
        "no_file": "Keine Datei",
        "processing": "Verarbeitung...",
        "q1": "Q1", "q2": "Q2", "q3": "Q3", "q4": "Q4",
        "half_year": "Halbjahr", "full_year": "Ganzes Jahr",
        "language": "Sprache",
        # --- YENİ EKLENEN ANAHTARLAR ---
        "bwa_history_title": "Verlauf der BWA-Uploads",
        "no_history": "Keine bisherigen Uploads",
        "delete_record_title": "Eintrag löschen",
        "confirm_delete_message": "Sind Sie sicher, dass Sie den Verlaufseintrag '{file_name}' dauerhaft löschen möchten?\n\nDieser Vorgang kann nicht rückgängig gemacht werden.",
        "record_not_found_error": "Eintrag konnte nicht gefunden und gelöscht werden.",
        "mapping_table_pos": "Pos.",
        "mapping_table_desc": "Beschreibung",
        "ai_status_message": "🤖 KI-Status: {status}",
        "bwa_loaded_from_history": "Verlauf geladen: {file_name}"
    },
    "TR": {
        "app_title": "EKS Form Doldurucu Pro",
        "customer": "Müşteri",
        "period": "Dönem",
        "template": "Şablon",
        "load_bwa": "BWA Dosyası Yükle",
        "auto_mapping": "Otomatik Eşleştirme",
        "export_eks": "EKS Dışa Aktar",
        "new_customer": "Yeni Müşteri",
        "customer_code": "Müşteri Kodu",
        "customer_name": "Müşteri Adı",
        "from_date": "Başlangıç Tarihi",
        "to_date": "Bitiş Tarihi",
        "quick_select": "Hızlı Seçim",
        "mapping_results": "Eşleştirme Sonuçları",
        "confidence": "Güven",
        "monthly_values": "Aylık Değerler",
        "total": "Toplam",
        "success": "Başarılı",
        "error": "Hata",
        "settings": "Ayarlar",
        "api_key": "API Anahtarı",
        "save": "Kaydet",
        "cancel": "İptal",
        "loading": "Yükleniyor...",
        "file_loaded": "Dosya yüklendi",
        "no_file": "Dosya yok",
        "processing": "İşleniyor...",
        "q1": "Q1", "q2": "Q2", "q3": "Q3", "q4": "Q4",
        "half_year": "6 Ay", "full_year": "12 Ay",
        "language": "Dil",
        # --- YENİ EKLENEN ANAHTARLAR ---
        "bwa_history_title": "Geçmiş BWA Yüklemeleri",
        "no_history": "Geçmiş yükleme yok",
        "delete_record_title": "Kaydı Sil",
        "confirm_delete_message": "'{file_name}' adlı geçmiş kaydını kalıcı olarak silmek istediğinizden emin misiniz?\n\nBu işlem geri alınamaz.",
        "record_not_found_error": "Kayıt bulunamadı ve silinemedi.",
        "mapping_table_pos": "Poz.",
        "mapping_table_desc": "Açıklama",
        "ai_status_message": "🤖 AI Durumu: {status}",
        "bwa_loaded_from_history": "Geçmişten yüklendi: {file_name}"
    }
}

@dataclass
class Customer:
    code: str
    name: str
    created_date: str
    default_template: str = "eks_standard.xlsx"
    notes: str = ""
    bwa_history: List[Dict] = field(default_factory=list) # Bu, dışa aktarım geçmişi
    bwa_upload_history: List[Dict] = field(default_factory=list) # Bu, YENİ yükleme geçmişi

@dataclass 
class MappingRule:
    eks_field: str
    bwa_source: str
    calculation_type: str  # 'direct', 'sum'
    source_accounts: List[str] = None
    description_de: str = ""

class ClaudeAPIHelper:
    """Claude API entegrasyonu için yardımcı sınıf"""
    
    def __init__(self, api_key: str = None):
        self.api_key = api_key
        self.base_url = "https://api.anthropic.com/v1/messages"
        
    def is_available(self) -> bool:
        """API kullanılabilir mi kontrol eder"""
        return bool(self.api_key and len(self.api_key) > 10)
    
    def suggest_mapping(self, account_code: str, description: str, bwa_context: str = "") -> Dict:
        """Bilinmeyen hesap kodu için EKS eşleştirme önerisi"""
        if not self.is_available():
            return {"suggestion": None, "confidence": 0, "reason": "API key not available"}
        
        try:
            prompt = f"""Du bist ein Experte für deutsche Buchführung und EKS-Formulare.

BWA Hesap Kodu: {account_code}
Beschreibung: {description}
Kontext: {bwa_context}

Welches EKS-Feld passt am besten zu diesem BWA-Konto? 

Verfügbare EKS-Felder:
A1: Betriebseinnahmen
A5: Vereinnahmte Umsatzsteuer  
A7: Vom Finanzamt erstattete Umsatzsteuer
B1: Wareneinkauf
B2c: Geringfügig Beschäftigte
B3: Raumkosten (Miete und Energiekosten)
B10: Büromaterial plus Porto
B11: Telefonkosten
B14c: Nebenkosten des Geldverkehrs
B14e: Reinigung
B14f: Repräsentationskosten
B14h: Sonstige Betriebliche Ausgaben
B17: Gezahlte Vorsteuer
B18: An Finanzamt gezahlte Umsatzsteuer

Antworte im JSON-Format:
{{"suggestion": "EKS_FIELD", "confidence": 85, "reason": "Kurze Begründung"}}"""

            headers = {
                "Content-Type": "application/json",
                "x-api-key": self.api_key,  # x-api-key kullan
                "anthropic-version": "2023-06-01"
            }
            
            data = {
                "model": "claude-3-haiku-20240307",  # Daha ucuz model
                "max_tokens": 200,
                "messages": [
                    {
                        "role": "user",
                        "content": prompt
                    }
                ]
            }
            
            response = requests.post(self.base_url, headers=headers, json=data, timeout=10)
            
            if response.status_code == 200:
                result = response.json()
                content = result["content"][0]["text"]
                
                # JSON'u düzgün parse et
                import re
                json_match = re.search(r'\{[^}]+\}', content)
                if json_match:
                    try:
                        suggestion = json.loads(json_match.group())
                        return suggestion
                    except json.JSONDecodeError:
                        return {"suggestion": None, "confidence": 0, "reason": "Invalid JSON response"}
                else:
                    return {"suggestion": None, "confidence": 0, "reason": "No JSON found in response"}
            else:
                error_msg = response.json().get('error', {}).get('message', f'Status: {response.status_code}')
                return {"suggestion": None, "confidence": 0, "reason": f"API Error: {error_msg}"}
                
        except requests.exceptions.Timeout:
            return {"suggestion": None, "confidence": 0, "reason": "Request timeout"}
        except Exception as e:
            return {"suggestion": None, "confidence": 0, "reason": f"Error: {str(e)}"}
        
class BWAParser:
    def __init__(self):
        self.mapping_rules = self._init_mapping_rules()
        self.bwa_data = None
        self.customer_info = None
        self.available_months = []
        self.claude_api = None  # Claude API helper
        
    def _init_mapping_rules(self) -> Dict[str, MappingRule]:
        return {
            # A Bölümü - Betriebseinnahmen
            "A1": MappingRule("A1", "Summe Erlöse", "direct", description_de="Betriebseinnahmen"),
            "A5": MappingRule("A5", "Summe Umsatzsteuer", "direct", description_de="Vereinnahmte Umsatzsteuer"),
            "A7": MappingRule("A7", "Ust-Erstattung", "direct", description_de="vom Finanzamt erstattete Umsatzsteuer"),
            
            # B Bölümü - Betriebsausgaben  
            "B1": MappingRule("B1", "Wareneinkauf", "sum", ["5400", "Summe Material, Stoffe, Waren"], "Wareneinkauf"),
            "B2c": MappingRule("B2c", "6030", "direct", ["6030", "6036", "6171"], "geringfügig Beschäftigte"),
            "B3": MappingRule("B3", "Miete + Energie", "sum", ["6310", "6325"], "Raumkosten (Miete und Energiekosten)"),
            "B11": MappingRule("B11", "6805", "direct", ["6805"], "Telefonkosten"),
            "B14c": MappingRule("B14c", "6855", "direct", ["6855"], "Nebenkosten des Geldverkehrs"),
            "B17": MappingRule("B17", "Summe Vorsteuer", "direct", description_de="gezahlte Vorsteuer"),
            
            # Ek mapping'ler
            "B10": MappingRule("B10", "Büromaterial", "sum", ["6815", "6800"], "Büromaterial plus Porto"),
            "B14e": MappingRule("B14e", "6330", "direct", ["6330"], "Reinigung"),
            "B14f": MappingRule("B14f", "6630", "direct", ["6630"], "Repräsentationskosten"),
            "B14h": MappingRule("B14h", "Sonstige", "sum", ["6300", "6850"], "sonst. Betriebliche Ausgaben"),
            "B18": MappingRule("B18", "3820", "direct", ["3820"], "an Finanzamt gezahlte USt")
        }
    
    def set_claude_api(self, api_key: str):
        """Claude API helper'ı ayarla"""
        self.claude_api = ClaudeAPIHelper(api_key)
        print(f"Claude API configured with key: {api_key[:20]}..." if len(api_key) > 20 else f"Claude API configured")
    
    def load_bwa_file(self, file_path: str) -> Tuple[bool, str]:
        try:
            # .xls ve .xlsx formatlarını destekle
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.xls':
                # Eski Excel formatı için xlrd kullan
                import xlrd
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)
                
                # xlrd verisini pandas DataFrame'e dönüştür
                data = []
                for row_idx in range(sheet.nrows):
                    data.append([sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)])
                df = pd.DataFrame(data)
            else:
                # Modern Excel formatı (.xlsx) - pandas ile direkt oku
                df = pd.read_excel(file_path, header=None, engine='openpyxl')
            
            print(f"\n{'='*60}")
            print(f"Excel loaded: {os.path.basename(file_path)}")
            print(f"Shape: {df.shape}")
            print(f"First 5 rows:")
            for i in range(min(5, len(df))):
                print(f"Row {i}: {df.iloc[i, 0]} | {df.iloc[i, 1] if df.shape[1] > 1 else ''}")
            print(f"{'='*60}\n")
            
            # Müşteri bilgisini bul - GELİŞTİRİLMİŞ
            customer_found = False
            for row_idx in range(min(10, len(df))):
                # Tüm sütunları kontrol et
                for col_idx in range(min(5, df.shape[1])):
                    try:
                        cell_value = df.iloc[row_idx, col_idx]
                        if pd.isna(cell_value):
                            continue
                        
                        cell_str = str(cell_value).strip()
                        
                        if len(cell_str) > 6:
                            # Boşlukla ayrılmış format: "111051 Sherzad Farman Jindi"
                            parts = cell_str.split(None, 1)
                            if len(parts) >= 2 and parts[0].isdigit() and len(parts[0]) >= 4:
                                self.customer_info = {
                                    "code": parts[0],
                                    "name": parts[1]
                                }
                                customer_found = True
                                print(f"Customer found at row {row_idx}, col {col_idx}: {self.customer_info}")
                                break
                            
                            # Köşeli parantez format: "[1105] Sherzad Farman Jindi"
                            if cell_str.startswith('[') and ']' in cell_str:
                                bracket_end = cell_str.index(']')
                                code_part = cell_str[1:bracket_end]
                                name_part = cell_str[bracket_end+1:].strip()
                                if code_part.isdigit() and name_part:
                                    self.customer_info = {
                                        "code": code_part,
                                        "name": name_part
                                    }
                                    customer_found = True
                                    print(f"Customer found (bracket) at row {row_idx}, col {col_idx}: {self.customer_info}")
                                    break
                    except Exception as e:
                        print(f"Error checking cell [{row_idx}, {col_idx}]: {e}")
                        continue
                
                if customer_found:
                    break
            
            if not customer_found:
                print("Warning: Customer info not detected")
            
            # Header satırını bul
            header_row = -1
            konto_col = -1
            bezeichnung_col = -1
            
            for row_idx in range(min(15, len(df))):
                row = df.iloc[row_idx]
                
                has_konto = False
                has_months = False
                
                for col_idx in range(len(row)):
                    try:
                        cell = row.iloc[col_idx]
                        if pd.isna(cell):
                            continue
                        
                        cell_str = str(cell).upper().strip()
                        
                        # "Konto" sütununu bul
                        if 'KONTO' in cell_str and konto_col == -1:
                            konto_col = col_idx
                            has_konto = True
                        
                        # "Bezeichnung" sütununu bul
                        if ('BEZEICHNUNG' in cell_str or 'DESCRIPTION' in cell_str) and bezeichnung_col == -1:
                            bezeichnung_col = col_idx
                        
                        # Ay kontrolü
                        if cell_str in ['JAN', 'FEB', 'MRZ', 'APR', 'MAI', 'JUN', 'JUL', 'AUG', 'SEP', 'OKT', 'NOV', 'DEZ']:
                            has_months = True
                    
                    except Exception as e:
                        continue
                
                if has_konto and has_months:
                    header_row = row_idx
                    print(f"Header found at row {header_row}")
                    print(f"  Konto col: {konto_col}, Bezeichnung col: {bezeichnung_col}")
                    break
            
            if header_row == -1:
                return False, "BWA Header nicht gefunden"
            
            # Veri satırlarını al
            data_start_row = header_row + 1
            self.bwa_data = df.iloc[data_start_row:].reset_index(drop=True)
            
            # Yeni DataFrame oluştur
            new_df = pd.DataFrame()
            
            # Konto ve Bezeichnung'u birleştir
            combined = []
            for idx in range(len(self.bwa_data)):
                try:
                    if konto_col >= 0 and bezeichnung_col >= 0 and konto_col != bezeichnung_col:
                        # FORMAT 2: Ayrı sütunlar
                        konto_val = self.bwa_data.iloc[idx, konto_col]
                        bez_val = self.bwa_data.iloc[idx, bezeichnung_col]
                        
                        konto_str = str(konto_val).strip() if pd.notna(konto_val) else ""
                        bez_str = str(bez_val).strip() if pd.notna(bez_val) else ""
                        
                        if konto_str == 'nan':
                            konto_str = ""
                        if bez_str == 'nan':
                            bez_str = ""
                        
                        combined_text = f"{konto_str} {bez_str}".strip()
                        combined.append(combined_text)
                    
                    elif konto_col >= 0:
                        # FORMAT 1: Tek sütun
                        val = self.bwa_data.iloc[idx, konto_col]
                        val_str = str(val).strip() if pd.notna(val) else ""
                        if val_str == 'nan':
                            val_str = ""
                        combined.append(val_str)
                    else:
                        combined.append("")
                
                except Exception as e:
                    print(f"Error at row {idx}: {e}")
                    combined.append("")
            
            new_df['Konto_Bezeichnung'] = combined
            
            # Ay sütunlarını ekle
            month_cols = ['JAN', 'FEB', 'MRZ', 'APR', 'MAI', 'JUN', 'JUL', 'AUG', 'SEP', 'OKT', 'NOV', 'DEZ']
            header_row_data = df.iloc[header_row]
            
            months_found = []
            for month in month_cols:
                for col_idx in range(len(header_row_data)):
                    try:
                        cell = header_row_data.iloc[col_idx]
                        if pd.notna(cell) and str(cell).upper().strip() == month:
                            new_df[month] = self.bwa_data.iloc[:, col_idx].values
                            months_found.append(month)
                            break
                    except Exception as e:
                        continue
            
            self.bwa_data = new_df
            self.available_months = months_found
            
            print(f"Available months: {self.available_months}")
            print(f"Final shape: {self.bwa_data.shape}")
            print(f"First 5 data rows:\n{self.bwa_data.head()}\n")
            
            if not self.available_months:
                return False, "Keine Monatsspalten gefunden"
            
            return True, f"BWA geladen: {len(self.available_months)} Monate verfügbar"
            
        except Exception as e:
            import traceback
            print("\nFULL ERROR:")
            traceback.print_exc()
            return False, f"Fehler beim Laden: {str(e)}"
    
    def extract_values_for_period(self, start_month: str, end_month: str) -> Dict:
        if self.bwa_data is None or self.bwa_data.empty:
            return {}
        
        # Monat-Indices bestimmen  
        month_order = ['JAN', 'FEB', 'MRZ', 'APR', 'MAI', 'JUN', 'JUL', 'AUG', 'SEP', 'OKT', 'NOV', 'DEZ']
        try:
            start_idx = month_order.index(start_month)
            end_idx = month_order.index(end_month)
            selected_months = month_order[start_idx:end_idx+1]
        except ValueError:
            selected_months = self.available_months[:6]
        
        # Nur verfügbare Monate verwenden
        selected_months = [m for m in selected_months if m in self.available_months]
        
        results = {}
        
        for field, rule in self.mapping_rules.items():
            extracted = self._extract_mapping(rule, selected_months)
            confidence = self._calculate_confidence(extracted['values'])
            
            results[field] = {
                'values': extracted['values'],
                'confidence': confidence,
                'source': rule.bwa_source,
                'description': rule.description_de,
                'months': selected_months,
                'total': sum(v for v in extracted['values'] if v is not None)
            }
        
        return results
    
    def _extract_mapping(self, rule: MappingRule, months: List[str]) -> Dict:
        if rule.calculation_type == "direct":
            return self._find_direct_match(rule.bwa_source, months)
        elif rule.calculation_type == "sum":
            return self._sum_multiple_accounts(rule.source_accounts, months)
        return {'values': [None] * len(months)}
    
    def _find_direct_match(self, search_term: str, months: List[str]) -> Dict:
        try:
            # İlk sütunu al (artık her zaman Konto+Bezeichnung kombinasyonu)
            if 'Konto_Bezeichnung' in self.bwa_data.columns:
                first_col = self.bwa_data['Konto_Bezeichnung'].astype(str)
            else:
                first_col = self.bwa_data.iloc[:, 0].astype(str)
            
            # Arama yap
            mask = first_col.str.contains(search_term, case=False, na=False, regex=False)
            
            if mask.any():
                row = self.bwa_data[mask].iloc[0]
                values = []
                for month in months:
                    if month in self.bwa_data.columns:
                        val = row[month]
                        values.append(float(val) if pd.notna(val) and val != '' else None)
                    else:
                        values.append(None)
                return {'values': values}
        except Exception as e:
            print(f"Error in _find_direct_match for '{search_term}': {e}")
        
        return {'values': [None] * len(months)}
    
    def _sum_multiple_accounts(self, accounts: List[str], months: List[str]) -> Dict:
        total_values = [0] * len(months)
        found_any = False
        
        for account in accounts:
            result = self._find_direct_match(account, months)
            for i, val in enumerate(result['values']):
                if val is not None:
                    total_values[i] += val
                    found_any = True
        
        return {'values': total_values if found_any else [None] * len(months)}
    
    def _calculate_confidence(self, values: List) -> int:
        if not values:
            return 0
        non_null = sum(1 for v in values if v is not None)
        return int((non_null / len(values)) * 100)
    
    def _find_unmapped_accounts(self) -> List[Dict]:
        """BWA'da bulunan ama mapping'de olmayan hesap kodlarını bulur - DÜZELTİLMİŞ"""
        if self.bwa_data is None or self.bwa_data.empty:
            return []
        
        # Mevcut mapping'deki tüm hesap kodlarını topla
        mapped_accounts = set()
        for rule in self.mapping_rules.values():
            if rule.source_accounts:
                mapped_accounts.update(rule.source_accounts)
            mapped_accounts.add(rule.bwa_source)
        
        unmapped = []
        try:
            first_col = self.bwa_data.iloc[:, 0].astype(str)
            second_col = self.bwa_data.iloc[:, 1].astype(str) if self.bwa_data.shape[1] > 1 else pd.Series()
            
            for idx in range(len(first_col)):
                text = first_col.iloc[idx]
                description = second_col.iloc[idx] if idx < len(second_col) else ""
                
                # 4-stellige Kontonummern veya önemli satırları bul
                account_match = re.match(r'^(\d{4})\s*(.*)$', text)
                if not account_match:
                    if text.isdigit() and len(text) == 4:
                        account_code = text
                        account_desc = description
                    else:
                        continue
                else:
                    account_code = account_match.group(1)
                    account_desc = account_match.group(2) if account_match.group(2) else description
                
                # Bu hesap zaten eşleştirilmiş mi?
                if account_code not in mapped_accounts:
                    # Bu satırda değerler var mı kontrol et
                    row_values = []
                    for month in self.available_months[:6]:
                        if month in self.bwa_data.columns:
                            val = self.bwa_data.iloc[idx][month]
                            if pd.notna(val) and val != '' and val != 0:
                                row_values.append(float(val))
                            else:
                                row_values.append(0)
                    
                    # Sadece değeri olan hesapları ekle
                    if any(v != 0 for v in row_values):
                        unmapped.append({
                            'account': account_code,
                            'description': account_desc[:100],
                            'values': row_values
                        })
                        print(f"Unmapped account found: {account_code} - {account_desc[:50]}")
        
        except Exception as e:
            print(f"Error finding unmapped accounts: {e}")
            import traceback
            traceback.print_exc()
        
        # En önemli 5 hesabı döndür
        unmapped.sort(key=lambda x: sum(abs(v) for v in x['values']), reverse=True)
        return unmapped[:5]
    
    def load_data_from_json(self, json_data: str, customer_info: Dict) -> Tuple[bool, str]:
        """Kaydedilmiş JSON verisinden BWA DataFrame'ini yeniden oluşturur."""
        try:
            # Kayıtlı JSON'dan DataFrame'i geri yükle
            self.bwa_data = pd.read_json(json_data, orient='split')
            self.customer_info = customer_info
            
            # Mevcut ayları yeniden hesapla
            month_cols = ['JAN', 'FEB', 'MRZ', 'APR', 'MAI', 'JUN', 'JUL', 'AUG', 'SEP', 'OKT', 'NOV', 'DEZ']
            self.available_months = [col for col in self.bwa_data.columns if col in month_cols]
            
            return True, f"BWA aus Verlauf geladen: {len(self.available_months)} Monate verfügbar"
        except Exception as e:
            self.bwa_data = None
            self.customer_info = None
            self.available_months = []
            return False, f"Fehler beim Laden aus Verlauf: {str(e)}"
    
    def _get_ai_suggestions(self, unmapped_accounts: List[Dict]) -> List[Dict]:
        """Claude API'den eşleştirme önerileri al - DÜZELTİLMİŞ"""
        if not self.claude_api or not self.claude_api.is_available():
            print("Claude API not available")
            return []
        
        suggestions = []
        
        for account in unmapped_accounts:
            print(f"Getting AI suggestion for account {account['account']}...")
            
            suggestion = self.claude_api.suggest_mapping(
                account['account'], 
                account['description'],
                f"Monatswerte: {account['values'][:3]}"
            )
            
            if suggestion.get('suggestion'):
                suggestions.append({
                    'bwa_account': account['account'],
                    'bwa_description': account['description'],
                    'suggested_eks': suggestion['suggestion'],
                    'confidence': suggestion.get('confidence', 0),
                    'reason': suggestion.get('reason', ''),
                    'values': account['values']
                })
                print(f"  -> Suggested: {suggestion['suggestion']} ({suggestion.get('confidence', 0)}%)")
            else:
                print(f"  -> No suggestion: {suggestion.get('reason', 'Unknown')}")
        
        return suggestions

class CustomerManager:
    def __init__(self, data_dir: str = "data"):
        self.data_dir = data_dir
        self.customers_dir = os.path.join(data_dir, "customers")
        os.makedirs(self.customers_dir, exist_ok=True)
        
    def save_customer(self, customer: Customer) -> bool:
        try:
            file_path = os.path.join(self.customers_dir, f"{customer.code}.json")
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(asdict(customer), f, ensure_ascii=False, indent=2)
            return True
        except Exception:
            return False
    
    def load_customer(self, customer_code: str) -> Optional[Customer]:
            try:
                file_path = os.path.join(self.customers_dir, f"{customer_code}.json")
                if os.path.exists(file_path):
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    # --- YENİ EKLENEN KONTROL ---
                    # Eski JSON dosyalarında bu alan olmayabilir, hata vermemesi için ekle
                    if 'bwa_upload_history' not in data:
                        data['bwa_upload_history'] = []
                    # --- KONTROL SONU ---

                    return Customer(**data)
            except Exception as e:
                # Hata ayıklama için print eklemek faydalı olabilir
                print(f"Error loading customer {customer_code}: {e}")
                pass
            return None
    
    def get_all_customers(self) -> List[Customer]:
        customers = []
        for file_name in os.listdir(self.customers_dir):
            if file_name.endswith('.json'):
                customer_code = file_name[:-5]
                customer = self.load_customer(customer_code)
                if customer:
                    customers.append(customer)
        return sorted(customers, key=lambda c: c.code)
    
class EKSFormFiller(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # Grundkonfiguration
        self.title("EKS Formular Ausfüller Pro")
        self.iconbitmap(resource_path("icon.ico"))
        self.geometry("1400x900")
        self.configure(fg_color="#1a1a1a")
        
        # Sprache
        self.language = "DE"
        self.texts = LANGUAGES[self.language]
        
        # Components
        self.bwa_parser = BWAParser()
        self.customer_manager = CustomerManager()
        
        # State
        self.current_customer = None
        self.bwa_file_path = None
        self.extracted_data = {}
        self.selected_start_month = "JAN"
        self.selected_end_month = "JUN"
        self.selected_year = datetime.now().year
        self.total_labels = {}
        
        # API Key'i yükle
        self.load_api_settings()
        
        self.setup_ui()
        self.load_customer_list()
    
    def load_api_settings(self):
        """API ayarlarını yükle - DÜZELTİLMİŞ"""
        try:
            settings_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "settings.json")
            if os.path.exists(settings_path):
                with open(settings_path, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    api_key = settings.get("claude_api_key", "")
                    if api_key:
                        print(f"Loading API key: {api_key[:20]}..." if len(api_key) > 20 else f"Loading API key")
                        self.bwa_parser.set_claude_api(api_key)
                        print("API key loaded successfully")
                    else:
                        print("No API key found in settings")
            else:
                print(f"Settings file not found at: {settings_path}")
        except Exception as e:
            print(f"API settings load error: {e}")
    
    def setup_ui(self):
        # Header
        header_frame = ctk.CTkFrame(self, height=80, fg_color="#2b2b2b")
        header_frame.pack(fill="x", padx=10, pady=5)
        header_frame.pack_propagate(False)

        # Titel und Controls in Header
        title_label = ctk.CTkLabel(header_frame, text=self.texts["app_title"], 
                                font=ctk.CTkFont(size=24, weight="bold"))
        title_label.pack(side="left", padx=20, pady=20)

        # Einstellungen Button
        settings_btn = ctk.CTkButton(header_frame, text="⚙️", width=40, height=40,
                                command=self.open_settings)
        settings_btn.pack(side="right", padx=10, pady=20)

        # --- YENİ EKLENEN VERSİYON ETİKETİ ---
        version_label = ctk.CTkLabel(header_frame, text=APP_VERSION, 
                                    font=ctk.CTkFont(size=11), text_color="gray")
        version_label.pack(side="right", padx=10, pady=5, anchor="s")
        # --- KOD SONU ---
        
        
        # Dil seçimi
        language_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        
        language_frame.pack(side="right", padx=10, pady=20)
        
        ctk.CTkLabel(language_frame, text=self.texts["language"] + ":", 
                    font=ctk.CTkFont(size=12)).pack(side="left", padx=5)
        
        self.language_combo = ctk.CTkComboBox(language_frame, values=["DE", "TR"], 
                                            width=60, command=self.change_language)
        self.language_combo.set(self.language)
        self.language_combo.pack(side="left", padx=5)
        
        # Main Container
        main_container = ctk.CTkFrame(self, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Control Panel (oben)
        control_frame = ctk.CTkFrame(main_container, height=120, fg_color="#2b2b2b")
        control_frame.pack(fill="x", pady=(0, 10))
        control_frame.pack_propagate(False)
        
        # Kunde Auswahl
        customer_frame = ctk.CTkFrame(control_frame, fg_color="#3b3b3b")
        customer_frame.pack(side="left", fill="y", padx=10, pady=10)
        
        ctk.CTkLabel(customer_frame, text=self.texts["customer"], 
                    font=ctk.CTkFont(weight="bold")).pack(pady=5)
        
        self.customer_combo = ctk.CTkComboBox(customer_frame, width=200, command=self.on_customer_selected)
        self.customer_combo.pack(pady=5)
        
        new_customer_btn = ctk.CTkButton(customer_frame, text="+", width=30, height=30,
                                        command=self.create_new_customer)
        new_customer_btn.pack(pady=5)
        
        # Zeitraum Auswahl
        period_frame = ctk.CTkFrame(control_frame, fg_color="#3b3b3b")
        period_frame.pack(side="left", fill="y", padx=10, pady=10)
        
        ctk.CTkLabel(period_frame, text=self.texts["period"], 
                    font=ctk.CTkFont(weight="bold")).pack(pady=5)
        
        # Yıl Seçimi
        year_frame = ctk.CTkFrame(period_frame, fg_color="transparent")
        year_frame.pack(pady=2)
        
        ctk.CTkLabel(year_frame, text="Jahr:", font=ctk.CTkFont(size=12)).pack(side="left", padx=2)
        years = [str(year) for year in range(2020, 2030)]
        self.year_combo = ctk.CTkComboBox(year_frame, values=years, width=80,
                                         command=self.on_year_changed)
        self.year_combo.set(str(self.selected_year))
        self.year_combo.pack(side="left", padx=2)
        
        # Ay Seçimi
        period_controls = ctk.CTkFrame(period_frame, fg_color="transparent")
        period_controls.pack(pady=5)
        
        months = ['JAN', 'FEB', 'MRZ', 'APR', 'MAI', 'JUN', 'JUL', 'AUG', 'SEP', 'OKT', 'NOV', 'DEZ']
        
        ctk.CTkLabel(period_controls, text="Von:", font=ctk.CTkFont(size=12)).pack(side="left", padx=2)
        self.start_month_combo = ctk.CTkComboBox(period_controls, values=months, width=60, 
                                               command=self.on_period_changed)
        self.start_month_combo.pack(side="left", padx=2)
        self.start_month_combo.set("JAN")
        
        ctk.CTkLabel(period_controls, text="Bis:", font=ctk.CTkFont(size=12)).pack(side="left", padx=2)
        
        self.end_month_combo = ctk.CTkComboBox(period_controls, values=months, width=60,
                                             command=self.on_period_changed)
        self.end_month_combo.pack(side="left", padx=2)
        self.end_month_combo.set("JUN")
        
        # Quick Select Buttons
        quick_frame = ctk.CTkFrame(period_frame, fg_color="transparent")
        quick_frame.pack(pady=5)
        
        quick_buttons = [
            ("Q1", lambda: self.set_period("JAN", "MRZ")),
            ("Q2", lambda: self.set_period("APR", "JUN")),
            ("Q3", lambda: self.set_period("JUL", "SEP")),
            ("Q4", lambda: self.set_period("OKT", "DEZ")),
        ]
        
        for text, command in quick_buttons:
            btn = ctk.CTkButton(quick_frame, text=text, width=35, height=25, command=command)
            btn.pack(side="left", padx=1)
        
        # Content Area (unten)
        content_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        content_frame.pack(fill="both", expand=True)
        
        # Sol panel - BWA Import
        left_panel = ctk.CTkFrame(content_frame, width=400, fg_color="#2b2b2b")
        left_panel.pack(side="left", fill="y", padx=(0, 10))
        left_panel.pack_propagate(False)
        
        ctk.CTkLabel(left_panel, text="BWA Import", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)
        
        self.load_bwa_btn = ctk.CTkButton(left_panel, text=self.texts["load_bwa"],
                                         command=self.load_bwa_file, height=40)
        self.load_bwa_btn.pack(pady=10, padx=20, fill="x")
        
        self.bwa_status_label = ctk.CTkLabel(left_panel, text=self.texts["no_file"], 
                                           text_color="gray")
        self.bwa_status_label.pack(pady=5)
        
        # BWA Info Anzeige
        self.bwa_info_frame = ctk.CTkFrame(left_panel, fg_color="#3b3b3b")
        self.bwa_info_frame.pack(fill="x", padx=20, pady=10)
        
        self.mapping_btn = ctk.CTkButton(left_panel, text=self.texts["auto_mapping"],
                                        command=self.perform_mapping, height=40, state="disabled")
        self.mapping_btn.pack(pady=20, padx=20, fill="x")
        
        # Template Analyse Button (Debug)
        analyze_btn = ctk.CTkButton(left_panel, text="🔍 Template Analysieren",
                                  command=self.analyze_template_wrapper, height=30)
        analyze_btn.pack(pady=5, padx=20, fill="x")
        
        self.export_btn = ctk.CTkButton(left_panel, text=self.texts["export_eks"],
                                       command=self.export_eks, height=40, state="disabled")
        self.export_btn.pack(pady=10, padx=20, fill="x")
                
        
        # --- YENİ ARAYÜZ BÖLÜMÜ BAŞLANGICI ---
        self.history_label = ctk.CTkLabel(left_panel, text=self.texts["bwa_history_title"], 
                                          font=ctk.CTkFont(size=14, weight="bold"))
        self.history_label.pack(pady=(20, 5), padx=20)
        
        self.bwa_history_frame = ctk.CTkScrollableFrame(left_panel, fg_color="#3b3b3b")
        self.bwa_history_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        # --- YENİ ARAYÜZ BÖLÜMÜ SONU ---

        # Rechts: Mapping Ergebnisse
        right_panel = ctk.CTkFrame(content_frame, fg_color="#2b2b2b")
        
        # Rechts: Mapping Ergebnisse
        right_panel = ctk.CTkFrame(content_frame, fg_color="#2b2b2b")
        right_panel.pack(side="right", fill="both", expand=True)
        
        ctk.CTkLabel(right_panel, text=self.texts["mapping_results"], 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)
        
        self.results_frame = ctk.CTkScrollableFrame(right_panel, fg_color="#1a1a1a")
        self.results_frame.pack(fill="both", expand=True, padx=20, pady=10)

    def on_year_changed(self, selected_year):
        """Yıl değiştiğinde çağrılır"""
        self.selected_year = int(selected_year)
    
    def set_period(self, start: str, end: str):
        self.start_month_combo.set(start)
        self.end_month_combo.set(end)
        self.on_period_changed()
    
    def on_period_changed(self, value=None):
        self.selected_start_month = self.start_month_combo.get()
        self.selected_end_month = self.end_month_combo.get()
    
    def change_language(self, selected_language):
        """Dil değiştirme fonksiyonu"""
        if selected_language == self.language:
            return
            
        self.language = selected_language
        self.texts = LANGUAGES[self.language]
        self.refresh_ui()
    
    def refresh_ui(self):
            """Dil değiştiğinde TÜM UI metinlerini günceller."""
            self.title(self.texts["app_title"])
            
            # Header ve Ana Butonlar
            # (Bu kısım zaten çalışıyor, ancak daha fazla eleman varsa buraya eklenebilir)
            self.load_bwa_btn.configure(text=self.texts["load_bwa"])
            self.mapping_btn.configure(text=self.texts["auto_mapping"])
            self.export_btn.configure(text=self.texts["export_eks"])
            
            # Sol Panel Başlıkları
            if hasattr(self, 'history_label'):
                self.history_label.configure(text=self.texts["bwa_history_title"])
            
            # Sağ Panel Başlığı
            # Bu başlığı yeniden çizmek daha kolay. Önce bir referans oluşturalım.
            # setup_ui içinde ilgili satırı `self.mapping_results_label = ctk.CTkLabel(...)` yapın.
            if hasattr(self, 'mapping_results_label'):
                self.mapping_results_label.configure(text=self.texts["mapping_results"])

            # Dinamik İçerikleri Yeniden Çiz
            # Bu, dil değişiminin her yerde görünür olmasını sağlar.
            self.display_bwa_history()
            self.display_mapping_results()
            
            # Durum Etiketleri
            current_text = self.bwa_status_label.cget("text")
            if "Keine Datei" in current_text or "Dosya yok" in current_text:
                self.bwa_status_label.configure(text=self.texts["no_file"])
            elif "geladen" in current_text or "yüklendi" in current_text or "loaded" in current_text:
                self.bwa_status_label.configure(text="✅ " + self.texts["file_loaded"])
    
    def load_customer_list(self):
        customers = self.customer_manager.get_all_customers()
        customer_options = [f"{c.code} - {c.name}" for c in customers]
        
        if customer_options:
            self.customer_combo.configure(values=customer_options)
            self.customer_combo.set(customer_options[0])
            # İlk müşteriyi otomatik seç
            self.on_customer_selected(customer_options[0])
            print(f"Loaded {len(customer_options)} customers")
        else:
            self.customer_combo.configure(values=["Keine Kunden"])
            self.current_customer = None
            print("No customers found")
    
    def on_customer_selected(self, selection):
        if selection and " - " in selection:
            customer_code = selection.split(" - ")[0]
            self.current_customer = self.customer_manager.load_customer(customer_code)
            
            if self.current_customer:
                print(f"Customer selected: {self.current_customer.code} - {self.current_customer.name}")
                self.display_bwa_history()
            else:
                print(f"Failed to load customer: {customer_code}")
                self.current_customer = None
        else:
            self.current_customer = None
            print("No valid customer selection")
    
    def create_new_customer(self):
        dialog = CustomerDialog(self, self.texts)
        if dialog.result:
            customer = Customer(
                code=dialog.result["code"],
                name=dialog.result["name"],
                created_date=datetime.now().strftime("%Y-%m-%d")
            )
            if self.customer_manager.save_customer(customer):
                self.load_customer_list()
                new_selection = f"{customer.code} - {customer.name}"
                self.customer_combo.set(new_selection)
                self.on_customer_selected(new_selection)
    
    def load_bwa_file(self):
        file_path = filedialog.askopenfilename(
            title="BWA Excel Datei auswählen",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("Excel 2007+", "*.xlsx"),
                ("Excel 97-2003", "*.xls"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.bwa_status_label.configure(text=self.texts["loading"])
            
            def load_thread():
                success, message = self.bwa_parser.load_bwa_file(file_path)
                self.after(0, lambda: self.on_bwa_loaded(success, message, file_path))
            
            threading.Thread(target=load_thread, daemon=True).start()
    
# form_doldurucu.py dosyasında

    def on_bwa_loaded(self, success: bool, message: str, file_path: str):
            if success:
                self.bwa_file_path = file_path
                self.bwa_status_label.configure(text="✅ " + self.texts["file_loaded"], text_color="green")
                self.mapping_btn.configure(state="normal")
                self.update_bwa_info()
                
                # BWA'dan gelen müşteri bilgisine göre işlem yap
                if self.bwa_parser.customer_info:
                    info = self.bwa_parser.customer_info
                    customer_code = info["code"]
                    
                    existing_customer = self.customer_manager.load_customer(customer_code)
                    
                    if not existing_customer:
                        self.auto_create_customer()
                    else:
                        self.current_customer = existing_customer
                        new_selection = f"{existing_customer.code} - {existing_customer.name}"
                        self.customer_combo.set(new_selection)

                # --- GÜNCELLENMİŞ KAYDETME BÖLÜMÜ ---
                # Yükleme geçmişini VERİ olarak kaydet
                if self.current_customer:
                    # pandas DataFrame'i JSON string'ine dönüştür
                    bwa_data_json = self.bwa_parser.bwa_data.to_json(orient='split')
                    
                    # Yeni geçmiş kaydını oluştur
                    new_entry = {
                        "date": datetime.now().strftime("%Y-%m-%d %H:%M"), # Daha okunaklı tarih
                        "file_name": os.path.basename(file_path),
                        "bwa_data_json": bwa_data_json,
                        "customer_info": self.bwa_parser.customer_info # Müşteri bilgisini de sakla
                    }
                    
                    # Müşterinin geçmiş listesine ekle
                    self.current_customer.bwa_upload_history.append(new_entry)
                    self.customer_manager.save_customer(self.current_customer)
                    self.display_bwa_history() # Geçmiş listesini yenile
                # --- GÜNCELLENMİŞ BÖLÜM SONU ---
                    
            else:
                self.bwa_status_label.configure(text="❌ " + message, text_color="red")


    def display_bwa_history(self):
            """Müşterinin geçmiş BWA yüklemelerini ve silme butonlarını arayüzde gösterir."""
            for widget in self.bwa_history_frame.winfo_children():
                widget.destroy()

            if self.current_customer and self.current_customer.bwa_upload_history:
                # Geçmişi tarihe göre ters sırala (en yeni en üstte)
                sorted_history = sorted(self.current_customer.bwa_upload_history, key=lambda x: x['date'], reverse=True)
                
                for entry in sorted_history[:10]: # Son 10 kaydı göster
                    
                    # Her kayıt için bir ana çerçeve oluştur
                    entry_frame = ctk.CTkFrame(self.bwa_history_frame, fg_color="#3b3b3b")
                    entry_frame.pack(fill="x", pady=2)

                    # Yükleme butonu (çerçevenin çoğunu kaplar)
                    btn_text = f"💾 {entry['date']} - {entry['file_name']}"
                    load_btn = ctk.CTkButton(
                        entry_frame,
                        text=btn_text,
                        anchor="w",
                        fg_color="transparent",
                        hover_color="#4a4a4a",
                        command=lambda e=entry: self.load_bwa_from_history(e)
                    )
                    load_btn.pack(side="left", fill="x", expand=True, padx=(5,0), pady=2)

                    # Silme butonu (sağda, küçük ve kırmızı)
                    delete_btn = ctk.CTkButton(
                        entry_frame,
                        text="❌",
                        width=30,
                        height=30,
                        fg_color="#c13e3e",
                        hover_color="#e05252",
                        font=ctk.CTkFont(size=14),
                        command=lambda e=entry: self.delete_bwa_history_entry(e)
                    )
                    delete_btn.pack(side="right", padx=5, pady=2)

            else:
                ctk.CTkLabel(self.bwa_history_frame, text=self.texts["no_history"]).pack(pady=10)

    def load_bwa_from_history(self, history_entry: Dict):
        """Geçmiş kayıttan bir BWA verisini yükler."""
        self.bwa_status_label.configure(text=self.texts["loading"])
        
        # Kaydedilmiş JSON verisini ve müşteri bilgisini al
        json_data = history_entry['bwa_data_json']
        customer_info = history_entry['customer_info']
        
        # Bu işlem çok hızlı olacağı için ayrı bir thread'e gerek yok
        success, message = self.bwa_parser.load_data_from_json(json_data, customer_info)
        
        if success:
            # Artık bir dosya yoluna bağlı değiliz
            self.bwa_file_path = None 
            self.bwa_status_label.configure(text=f'✅ {self.texts["bwa_loaded_from_history"].format(file_name=history_entry["file_name"])}', text_color="green")
            self.mapping_btn.configure(state="normal")
            self.update_bwa_info()
        else:
            self.bwa_status_label.configure(text="❌ " + message, text_color="red")
            self.mapping_btn.configure(state="disabled")

    def delete_bwa_history_entry(self, entry_to_delete: Dict):
            """Seçilen bir geçmiş BWA kaydını kullanıcı onayıyla siler."""
            if not self.current_customer:
                return

            file_name = entry_to_delete.get('file_name', 'Bilinmeyen Kayıt')
            
            # Kullanıcıdan onay al (Dinamik metinlerle)
            confirm_message = self.texts["confirm_delete_message"].format(file_name=file_name)
            confirm = messagebox.askyesno(
                self.texts["delete_record_title"],
                confirm_message,
                icon='warning'
            )

            if confirm:
                try:
                    self.current_customer.bwa_upload_history.remove(entry_to_delete)
                    self.customer_manager.save_customer(self.current_customer)
                    self.display_bwa_history()
                except ValueError:
                    messagebox.showerror(self.texts["error"], self.texts["record_not_found_error"])
    
    def update_bwa_info(self):
        for widget in self.bwa_info_frame.winfo_children():
            widget.destroy()
        
        if self.bwa_parser.customer_info:
            info = self.bwa_parser.customer_info
            ctk.CTkLabel(self.bwa_info_frame, text=f"Kunde: {info['code']}", 
                        font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=2)
            ctk.CTkLabel(self.bwa_info_frame, text=f"Name: {info['name']}").pack(anchor="w", padx=10, pady=2)
        
        if self.bwa_parser.available_months:
            months_text = ", ".join(self.bwa_parser.available_months)
            ctk.CTkLabel(self.bwa_info_frame, text=f"Monate: {months_text}").pack(anchor="w", padx=10, pady=2)
    
    def auto_create_customer(self):
        if self.bwa_parser.customer_info:
            info = self.bwa_parser.customer_info
            existing = self.customer_manager.load_customer(info["code"])
            if not existing:
                customer = Customer(
                    code=info["code"],
                    name=info["name"],
                    created_date=datetime.now().strftime("%Y-%m-%d")
                )
                if self.customer_manager.save_customer(customer):
                    self.load_customer_list()
                    new_selection = f"{customer.code} - {customer.name}"
                    self.customer_combo.set(new_selection)
                    self.on_customer_selected(new_selection)
    
    def perform_mapping(self):
            """DÜZELTİLMİŞ perform_mapping fonksiyonu"""
            if self.bwa_parser.bwa_data is None or self.bwa_parser.bwa_data.empty:
                return
            
            self.mapping_btn.configure(text=self.texts["processing"], state="disabled")
            
            # Progress göstergesi
            progress_window = ctk.CTkToplevel(self)
            progress_window.title("Verarbeitung...")
            progress_window.geometry("300x100")
            progress_window.transient(self)
            progress_window.grab_set()
            
            # Progress window'u merkeze al
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (150)
            y = (progress_window.winfo_screenheight() // 2) - (50)
            progress_window.geometry(f"300x100+{x}+{y}")
            
            progress_label = ctk.CTkLabel(progress_window, text="Daten werden verarbeitet...", 
                                        font=ctk.CTkFont(size=14))
            progress_label.pack(pady=20)
            
            progress_bar = ctk.CTkProgressBar(progress_window)
            progress_bar.pack(pady=10, padx=20, fill="x")
            progress_bar.set(0.3)
            
            def mapping_thread():
                try:
                    # Temel eşleştirme
                    extracted = self.bwa_parser.extract_values_for_period(
                        self.selected_start_month, self.selected_end_month
                    )
                    
                    # Progress güncelle
                    self.after(0, lambda: progress_bar.set(0.6))
                    self.after(0, lambda: progress_label.configure(text="Claude AI Vorschläge werden abgerufen..."))
                    
                    # --- İYİLEŞTİRİLMİŞ BÖLÜM BAŞLANGICI ---
                    # Claude API aktifse öneriler al
                    if self.bwa_parser.claude_api and self.bwa_parser.claude_api.is_available():
                        unmapped = self.bwa_parser._find_unmapped_accounts()
                        if unmapped:
                            print(f"Found {len(unmapped)} unmapped accounts, getting AI suggestions...")
                            ai_suggestions = self.bwa_parser._get_ai_suggestions(unmapped)
                            if ai_suggestions:
                                extracted['_ai_suggestions'] = ai_suggestions
                                print(f"Got {len(ai_suggestions)} AI suggestions")
                            else:
                                # AI'dan öneri gelmediyse (geçersiz anahtar vb.) durumu not et
                                extracted['_ai_status'] = "AI önerileri alınamadı. API anahtarı geçersiz olabilir."
                                print("No AI suggestions received (API key may be invalid)")
                        else:
                            # Eşleştirilecek yeni hesap bulunamadıysa durumu not et
                            extracted['_ai_status'] = "Tüm hesaplar eşleştirilmiş görünüyor."
                    else:
                        # API hiç yapılandırılmadıysa durumu not et
                        extracted['_ai_status'] = "Claude AI aktif değil. Ayarlardan API anahtarınızı girin."
                        print("Claude API not configured or not available")
                    # --- İYİLEŞTİRİLMİŞ BÖLÜM SONU ---
                    
                    # Progress tamamlandı
                    self.after(0, lambda: progress_bar.set(1.0))
                    self.after(0, lambda: progress_window.destroy())
                    
                    self.after(0, lambda: self.handle_mapping_complete(extracted))
                    
                except Exception as e:
                    print(f"Mapping error: {e}")
                    import traceback
                    traceback.print_exc()
                    self.after(0, lambda: progress_window.destroy())
                    self.after(0, lambda: messagebox.showerror("Fehler", f"Mapping fehlgeschlagen: {str(e)}"))
                    self.after(0, lambda: self.mapping_btn.configure(text=self.texts["auto_mapping"], state="normal"))
            
            threading.Thread(target=mapping_thread, daemon=True).start()
    
    def handle_mapping_complete(self, extracted_data: Dict):
        """Mapping tamamlandığında çağrılır"""
        self.extracted_data = extracted_data
        self.mapping_btn.configure(text=self.texts["auto_mapping"], state="normal")
        
        # Sadece gerçek veri varsa export butonunu aktifleştir
        has_real_data = any(not key.startswith('_') for key in extracted_data.keys())
        
        if has_real_data:
            self.export_btn.configure(state="normal")
            print("Export button enabled - data available")
        else:
            self.export_btn.configure(state="disabled")
            print("Export button disabled - no valid data")
        
        self.display_mapping_results()


    def display_mapping_results(self):
            # Önceki sonuçları ve etiket referanslarını temizle
            for widget in self.results_frame.winfo_children():
                widget.destroy()
            self.total_labels = {}
            
            if not self.extracted_data or not any(self.extracted_data.values()):
                ctk.CTkLabel(self.results_frame, text="Keine Ergebnisse").pack(pady=20)
                return

            # Sütun ağırlıklarını ayarla
            self.results_frame.grid_columnconfigure(0, weight=1, minsize=50)   # Pos.
            self.results_frame.grid_columnconfigure(1, weight=5, minsize=250)  # Beschreibung
            
            months = list(next(iter(self.extracted_data.values())).get('months', []))
            num_month_cols = len(months)
            
            for i in range(num_month_cols):
                self.results_frame.grid_columnconfigure(i + 2, weight=2, minsize=100) # Aylık Sütunlar
            self.results_frame.grid_columnconfigure(num_month_cols + 2, weight=3, minsize=120) # Gesamt

        # === Başlık Satırı (Grid ile mükemmel hizalama) ===
            header_font = ctk.CTkFont(size=12, weight="bold")
            header_fg_color = "#333333"
            
            ctk.CTkLabel(self.results_frame, text=self.texts["mapping_table_pos"], font=header_font, fg_color=header_fg_color, corner_radius=0, anchor="w").grid(row=0, column=0, sticky="ew", padx=(0,1), pady=(0,1))
            ctk.CTkLabel(self.results_frame, text=self.texts["mapping_table_desc"], font=header_font, fg_color=header_fg_color, corner_radius=0, anchor="w").grid(row=0, column=1, sticky="ew", padx=(0,1), pady=(0,1))
            for i, month in enumerate(months):
                ctk.CTkLabel(self.results_frame, text=month, font=header_font, fg_color=header_fg_color, corner_radius=0).grid(row=0, column=i + 2, sticky="ew", padx=(0,1), pady=(0,1))
            ctk.CTkLabel(self.results_frame, text=self.texts["total"], font=header_font, fg_color=header_fg_color, corner_radius=0).grid(row=0, column=num_month_cols + 2, sticky="ew", padx=(0,1), pady=(0,1))

            # === Veri Satırları ===
            row_index = 1
            sorted_fields = sorted([k for k in self.extracted_data.keys() if not k.startswith('_')])

            for field in sorted_fields:
                data = self.extracted_data[field]
                
                # EKS Kodu ve Açıklama
                ctk.CTkLabel(self.results_frame, text=field, font=ctk.CTkFont(weight="bold"), anchor="w").grid(row=row_index, column=0, sticky="w", padx=10, pady=8)
                ctk.CTkLabel(self.results_frame, text=data['description'], anchor="w").grid(row=row_index, column=1, sticky="ew", padx=5)

                # Aylık Değerler (Düzenlenebilir)
                for i, value in enumerate(data['values']):
                    EditableLabel(self.results_frame, row_index, i + 2, value, self.update_data_value)

                # Toplam
                total = data.get('total', 0)
                self.total_labels[field] = EditableLabel(self.results_frame, row_index, num_month_cols + 2, total, lambda r,c,v: None, is_total=True)
                
                # Ayırıcı Çizgi
                separator = ctk.CTkFrame(self.results_frame, height=1, fg_color="#3a3a3a")
                separator.grid(row=row_index + 1, column=0, columnspan=num_month_cols + 3, sticky="ew")
                
                row_index += 2 # Ayırıcı için ekstra satır atla
                
            # === AI Önerileri (tablonun altında) ===
            row_index += 1
            if '_ai_suggestions' in self.extracted_data and self.extracted_data['_ai_suggestions']:
                ai_frame = ctk.CTkFrame(self.results_frame, fg_color="transparent")
                ai_frame.grid(row=row_index, column=0, columnspan=num_month_cols + 3, sticky="ew")
                self.display_ai_suggestions(self.extracted_data['_ai_suggestions']) # display_ai_suggestions'ı bu çerçeveye çizecek şekilde ayarlamak gerekebilir
            elif '_ai_status' in self.extracted_data:
                ai_status_frame = ctk.CTkFrame(self.results_frame, fg_color="#4a4a4a")
                ai_status_frame.grid(row=row_index, column=0, columnspan=num_month_cols + 3, sticky="ew", pady=20, padx=10)
                status_text = self.texts["ai_status_message"].format(status=self.extracted_data['_ai_status']) # Dinamik metin
                ctk.CTkLabel(ai_status_frame, text=status_text, font=ctk.CTkFont(size=12)).pack(pady=10, padx=10)
    
    def display_ai_suggestions(self, suggestions: List[Dict]):
        """Claude AI önerilerini gösterir"""
        if not suggestions:
            return
        
        # AI Suggestions başlığı
        ai_header = ctk.CTkFrame(self.results_frame, fg_color="#1a4d1a")
        ai_header.pack(fill="x", pady=10, padx=10)
        
        ctk.CTkLabel(ai_header, text="🤖 Claude AI Vorschläge", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)
        
        for suggestion in suggestions:
            suggestion_frame = ctk.CTkFrame(self.results_frame, fg_color="#2d4a2d")
            suggestion_frame.pack(fill="x", pady=5, padx=10)
            
            # Header
            header_text = f"BWA {suggestion['bwa_account']}: {suggestion['bwa_description'][:50]}..."
            header_label = ctk.CTkLabel(suggestion_frame, text=header_text, 
                                      font=ctk.CTkFont(weight="bold"))
            header_label.pack(anchor="w", padx=10, pady=5)
            
            # Vorschlag
            suggestion_text = f"➜ {suggestion['suggested_eks']} (Vertrauen: {suggestion['confidence']}%)"
            suggestion_label = ctk.CTkLabel(suggestion_frame, text=suggestion_text, 
                                          text_color="#90EE90")
            suggestion_label.pack(anchor="w", padx=20, pady=2)
            
            # Begründung
            if suggestion['reason']:
                reason_label = ctk.CTkLabel(suggestion_frame, text=f"Grund: {suggestion['reason']}", 
                                          font=ctk.CTkFont(size=11))
                reason_label.pack(anchor="w", padx=20, pady=2)
            
            # Werte
            values_text = " | ".join([f"{val:.0f}" if val != 0 else "0" for val in suggestion['values']])
            values_label = ctk.CTkLabel(suggestion_frame, text=f"Werte: {values_text}", 
                                      font=ctk.CTkFont(size=10))
            values_label.pack(anchor="w", padx=20, pady=2)
            
            # Action Button
            button_frame = ctk.CTkFrame(suggestion_frame, fg_color="transparent")
            button_frame.pack(anchor="w", padx=20, pady=5)
            
            accept_btn = ctk.CTkButton(button_frame, text="✓ Akzeptieren", width=100, height=25,
                                     command=lambda s=suggestion: self.accept_ai_suggestion(s))
            accept_btn.pack(side="left", padx=5)
            
            ignore_btn = ctk.CTkButton(button_frame, text="✗ Ignorieren", width=100, height=25,
                                     fg_color="gray", command=lambda s=suggestion: self.ignore_ai_suggestion(s))
            ignore_btn.pack(side="left", padx=5)
    
    def accept_ai_suggestion(self, suggestion: Dict):
        """AI önerisini kabul et"""
        try:
            new_rule = MappingRule(
                suggestion['suggested_eks'],
                suggestion['bwa_account'],
                "direct",
                [suggestion['bwa_account']],
                suggestion['bwa_description'][:30]
            )
            
            self.bwa_parser.mapping_rules[suggestion['suggested_eks']] = new_rule
            self.perform_mapping()
            
            messagebox.showinfo("AI Vorschlag", 
                f"Zuordnung {suggestion['bwa_account']} → {suggestion['suggested_eks']} wurde hinzugefügt!")
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Hinzufügen der Zuordnung: {e}")
    
    def ignore_ai_suggestion(self, suggestion: Dict):
        """AI önerisini görmezden gel"""
        pass
    
    def open_settings(self):
        settings_dialog = SettingsDialog(self, self.texts)
    

    def analyze_template_wrapper(self):
        """Template analiz fonksiyonunu çağırır"""
        analysis = self.analyze_template_structure()
        if analysis:
            result_text = f"""Template Analizi Tamamlandı!

Müşteri Alanları: {len(analysis['customer_fields'])} adet
Ay Sütunları: {len(analysis['month_columns'])} adet  
EKS Pozisyonları: {len(analysis['data_positions'])} adet

Konsol çıktısını kontrol edin."""
            
            messagebox.showinfo("Template Analizi", result_text)
        else:
            messagebox.showerror("Hata", "Template analizi başarısız. templates/eks_form.xlsx dosyası var mı?")
    
    def analyze_template_structure(self):
        """Template yapısını analiz eder"""
        template_path = os.path.join("templates", "eks_form.xlsx")
        if not os.path.exists(template_path):
            return None
            
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            analysis = {
                "customer_fields": {},
                "data_positions": {},
                "month_columns": [],
                "structure": []
            }
            
            print("=== EKS TEMPLATE ANALYSE ===")
            
            for row in range(1, min(100, ws.max_row + 1)):
                for col in range(1, min(20, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        cell_text = str(cell.value).strip()
                        col_letter = chr(ord('A') + col - 1)
                        
                        # Müşteri bilgi alanlarını bul
                        if "Nummer der Bedarfsgemeinschaft" in cell_text:
                            analysis["customer_fields"]["number"] = f"{col_letter}{row}"
                        elif "Name, Vorname" in cell_text:
                            analysis["customer_fields"]["name"] = f"{col_letter}{row}"
                        elif "Bewilligungszeitraum" in cell_text:
                            analysis["customer_fields"]["period"] = f"{col_letter}{row}"
                        
                        # Ay başlıklarını bul
                        if cell_text in ['JAN', 'FEB', 'MRZ', 'APR', 'MAI', 'JUN', 'JUL', 'AUG', 'SEP', 'OKT', 'NOV', 'DEZ']:
                            analysis["month_columns"].append((cell_text, col))
                        
                        # EKS kodlarını bul
                        if len(cell_text) <= 4 and any(cell_text.startswith(prefix) for prefix in ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'B1', 'B2', 'B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B9', 'B10', 'B11', 'B12', 'B13', 'B14', 'B15', 'B16', 'B17', 'B18']):
                            analysis["data_positions"][cell_text] = {"row": row, "col": col}
                        
                        if row <= 50:
                            analysis["structure"].append({
                                "position": f"{col_letter}{row}",
                                "content": cell_text[:50] + "..." if len(cell_text) > 50 else cell_text
                            })
            
            print("Müşteri Alanları:", analysis["customer_fields"])
            print("Ay Sütunları:", analysis["month_columns"])
            print("EKS Pozisyonları:", analysis["data_positions"])
            
            return analysis
            
        except Exception as e:
            print(f"Template analysis error: {e}")
            return None
    
    def export_eks(self):
        # 1. Müşteri kontrolü
        if not self.current_customer:
            warning_msg = "Kein Kunde ausgewählt. Bitte wählen Sie zuerst einen Kunden aus!" if self.language == "DE" else "Müşteri seçilmedi. Lütfen önce bir müşteri seçin!"
            messagebox.showwarning("Warnung" if self.language == "DE" else "Uyarı", warning_msg)
            return
        
        # 2. Veri kontrolü
        if not self.extracted_data:
            warning_msg = "Keine Daten zum Exportieren. Bitte zuerst 'Automatische Zuordnung' ausführen!" if self.language == "DE" else "Dışa aktarılacak veri yok. Lütfen önce 'Otomatik Eşleştirme' yapın!"
            messagebox.showwarning("Warnung" if self.language == "DE" else "Uyarı", warning_msg)
            return
        
        # 3. Gerçek veri var mı kontrol et
        has_real_data = any(not key.startswith('_') for key in self.extracted_data.keys())
        if not has_real_data:
            warning_msg = "Keine gültigen Daten zum Exportieren gefunden!" if self.language == "DE" else "Geçerli veri bulunamadı!"
            messagebox.showwarning("Warnung" if self.language == "DE" else "Uyarı", warning_msg)
            return
        
        template_dir = "templates"
        os.makedirs(template_dir, exist_ok=True)
        
        try:
            # Dosya adı oluştur
            filename = f"{self.current_customer.code}_EKS_{self.selected_start_month}-{self.selected_end_month}_{self.selected_year}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            # Kayıt yeri seç
            export_path = filedialog.asksaveasfilename(
                title="EKS Export speichern" if self.language == "DE" else "EKS Dışa Aktar",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=filename
            )
            
            # Kullanıcı iptal ettiyse
            if not export_path:
                return
            
            # Export işlemini gerçekleştir
            success = self.create_eks_export(export_path)
            
            if success:
                success_msg = f"EKS erfolgreich exportiert:\n{export_path}" if self.language == "DE" else f"EKS başarıyla dışa aktarıldı:\n{export_path}"
                messagebox.showinfo("Erfolg" if self.language == "DE" else "Başarılı", success_msg)
                self.update_customer_history()
            else:
                error_msg = "Export fehlgeschlagen. Bitte Template-Datei überprüfen." if self.language == "DE" else "Dışa aktarma başarısız. Lütfen şablon dosyasını kontrol edin."
                messagebox.showerror("Fehler" if self.language == "DE" else "Hata", error_msg)
        
        except Exception as e:
            import traceback
            traceback.print_exc()
            error_msg = f"Export Fehler: {str(e)}" if self.language == "DE" else f"Dışa Aktarma Hatası: {str(e)}"
            messagebox.showerror("Fehler" if self.language == "DE" else "Hata", error_msg)
    
    def create_eks_export(self, export_path: str) -> bool:
            """
            Elde edilen verileri kullanarak EKS Excel dosyasını oluşturur.
            Bu versiyon, Excel şablonunu doğrudan kodun içine gömülü veriden alır.
            """
            try:
                # --- YENİ VE GARANTİLİ YÖNTEM BAŞLANGICI ---

                # Adım 1: Koda gömülü metin verisini al ve çöz.
                # 'template_data.py' dosyasındaki 'b64_data' değişkenini kullanıyoruz.
                # base64.b64decode, bu uzun metni tekrar orijinal Excel dosyasının
                # ikili (binary) verisine dönüştürür.
                template_content = base64.b64decode(template_data.b64_data)

                # Adım 2: Güvenli, geçici bir Excel dosyası oluştur.
                # openpyxl kütüphanesi bir dosya yoluyla çalışmak zorundadır,
                # bu yüzden çözdüğümüz bu ikili veriyi geçici bir dosyaya yazıyoruz.
                # 'delete=False' önemlidir, çünkü bu, dosyayı biz silene kadar tutar.
                # 'suffix='.xlsx'' ise dosyanın uzantısının doğru olmasını sağlar.
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                    temp_file.write(template_content)
                    # Geçici dosyanın tam yolunu bir değişkene atıyoruz.
                    # Örn: C:\Users\Kullanıcı\AppData\Local\Temp\tmp123abc.xlsx
                    template_path = temp_file.name
                
                # --- YÖNTEM SONU ---

                # Adım 3: Geçici şablon dosyasını openpyxl ile aç.
                # Kodun geri kalanı için her şey eskisi gibi. openpyxl, bunun geçici
                # bir dosya olduğunu bilmez, normal bir Excel dosyası gibi davranır.
                wb = openpyxl.load_workbook(template_path)
                ws = wb.active
                
                # Adım 4: Formu doldurma işlemlerini çağır.
                # Bu fonksiyonlar değişmedi, aynı şekilde çalışmaya devam ediyorlar.
                success = self.fill_eks_template(ws)
                if not success:
                    os.remove(template_path) # Hata olursa bile geçici dosyayı sil
                    return False
                
                self.update_customer_info_in_template(ws)
                self.update_period_info_in_template(ws)
                
                # Adım 5: Doldurulmuş son halini kullanıcının istediği yere kaydet.
                wb.save(export_path)
                
                # Adım 6: Temizlik.
                # Artık işimiz bittiğine göre, oluşturduğumuz geçici dosyayı
                # sistemden kalıcı olarak siliyoruz. Bu, gereksiz dosya birikimini önler.
                os.remove(template_path)
                
                # Her şey başarılı olduysa True döndür.
                return True
                
            except Exception as e:
                # Herhangi bir hata olursa konsola yazdır ve False döndür.
                print(f"Template Export Hatası: {e}")
                # Eğer hata sırasında geçici dosya hala varsa, onu silmeye çalış.
                if 'template_path' in locals() and os.path.exists(template_path):
                    os.remove(template_path)
                return False
    
    def fill_eks_template(self, ws) -> bool:
        """EKS template'indeki hücreleri doldurur"""
        try:
            eks_positions = {
                # A Bölümü - Betriebseinnahmen (Satır 10-17)
                "A1": {"start_row": 10, "months_start_col": 3},
                "A2": {"start_row": 11, "months_start_col": 3},
                "A3": {"start_row": 12, "months_start_col": 3},
                "A4": {"start_row": 13, "months_start_col": 3},
                "A5": {"start_row": 14, "months_start_col": 3},
                "A6": {"start_row": 15, "months_start_col": 3},
                "A7": {"start_row": 16, "months_start_col": 3},
                
                # B Bölümü - Betriebsausgaben (Satır 22-67)
                "B1": {"start_row": 22, "months_start_col": 3},
                "B2a": {"start_row": 24, "months_start_col": 3},
                "B2b": {"start_row": 25, "months_start_col": 3},
                "B2c": {"start_row": 26, "months_start_col": 3},
                "B2d": {"start_row": 27, "months_start_col": 3},
                "B3": {"start_row": 28, "months_start_col": 3},
                "B4": {"start_row": 29, "months_start_col": 3},
                "B5": {"start_row": 30, "months_start_col": 3},
                "B5_1a": {"start_row": 33, "months_start_col": 3},
                "B5_1b": {"start_row": 34, "months_start_col": 3},
                "B5_1c": {"start_row": 35, "months_start_col": 3},
                "B5_1d": {"start_row": 36, "months_start_col": 3},
                "B10": {"start_row": 50, "months_start_col": 3},
                "B11": {"start_row": 51, "months_start_col": 3},
                "B12": {"start_row": 52, "months_start_col": 3},
                "B14c": {"start_row": 57, "months_start_col": 3},
                "B14e": {"start_row": 59, "months_start_col": 3},
                "B14f": {"start_row": 60, "months_start_col": 3},
                "B14h": {"start_row": 62, "months_start_col": 3},
                "B17": {"start_row": 66, "months_start_col": 3},
                "B18": {"start_row": 67, "months_start_col": 3}
            }
            
            for field, data in self.extracted_data.items():
                if field.startswith('_'):
                    continue
                    
                if field in eks_positions:
                    pos = eks_positions[field]
                    row = pos["start_row"]
                    start_col = pos["months_start_col"]
                    
                    values = data.get('values', [])
                    for i, value in enumerate(values):
                        if value is not None and i < 6:
                            col = start_col + i
                            col_letter = chr(ord('A') + col - 1)
                            ws[f'{col_letter}{row}'] = value
                            ws[f'{col_letter}{row}'].number_format = '#,##0.00'
                    
                    print(f"Filled {field} at row {row}: {values}")
            
            return True
            
        except Exception as e:
            print(f"Template fill error: {e}")
            return False
    
    def update_customer_info_in_template(self, ws):
        """Müşteri bilgilerini template'e yazar"""
        try:
            if self.current_customer:
                ws['D2'] = self.current_customer.code
                ws['D3'] = self.current_customer.name
        except Exception as e:
            print(f"Customer info update error: {e}")
    
    def update_period_info_in_template(self, ws):
        """Dönem bilgilerini template'e yazar"""
        try:
            months = list(self.extracted_data.values())[0].get('months', [])
            if not months:
                return
                
            month_to_number = {
                'JAN': '01', 'FEB': '02', 'MRZ': '03', 'APR': '04', 
                'MAI': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
                'SEP': '09', 'OKT': '10', 'NOV': '11', 'DEZ': '12'
            }
            
            start_month_num = month_to_number.get(months[0], '01')
            end_month_num = month_to_number.get(months[-1], '06')
            selected_year = self.selected_year
            
            for row in range(1, 20):
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and "Bewilligungszeitraum vom" in str(cell.value):
                        original_text = str(cell.value)
                        updated_text = original_text.replace(
                            "_01.0x.200x__", f"01.{start_month_num}.{selected_year}"
                        ).replace(
                            "_3x.0x.200x__", f"30.{end_month_num}.{selected_year}"
                        )
                        cell.value = updated_text
                        print(f"Period updated: {updated_text}")
                        break
                            
        except Exception as e:
            print(f"Period info update error: {e}")
    
    def create_automatic_export(self, export_path: str) -> bool:
        """Fallback: Otomatik template oluşturur"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "EKS Formular"
            
            header_font = Font(bold=True, size=12)
            title_font = Font(bold=True, size=14)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            ws['A1'] = "Angaben zum voraussichtlichen Einkommen aus selbständiger Tätigkeit"
            ws['A1'].font = title_font
            ws.merge_cells('A1:H1')
            
            ws['A3'] = f"Nummer der Bedarfsgemeinschaft: {self.current_customer.code}"
            ws['A4'] = f"Name, Vorname: {self.current_customer.name}"
            ws['A5'] = f"Bewilligungszeitraum: {self.selected_start_month} - {self.selected_end_month} {self.selected_year}"
            
            months = list(self.extracted_data.values())[0].get('months', [])
            row = 8
            
            ws[f'A{row}'] = "Position"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            
            ws[f'B{row}'] = "Beschreibung"
            ws[f'B{row}'].font = header_font
            ws[f'B{row}'].fill = header_fill
            
            col_start = 3
            for i, month in enumerate(months):
                col_letter = chr(ord('C') + i)
                ws[f'{col_letter}{row}'] = month
                ws[f'{col_letter}{row}'].font = header_font
                ws[f'{col_letter}{row}'].fill = header_fill
            
            sum_col = chr(ord('C') + len(months))
            ws[f'{sum_col}{row}'] = "Summe"
            ws[f'{sum_col}{row}'].font = header_font
            ws[f'{sum_col}{row}'].fill = header_fill
            
            current_row = row + 1
            
            # A. Betriebseinnahmen
            ws[f'A{current_row}'] = "A. Betriebseinnahmen"
            ws[f'A{current_row}'].font = header_font
            current_row += 1
            
            for field, data in self.extracted_data.items():
                if field.startswith('_'):
                    continue
                if field.startswith('A'):
                    ws[f'A{current_row}'] = field
                    ws[f'B{current_row}'] = data['description']
                    
                    for i, value in enumerate(data['values']):
                        col_letter = chr(ord('C') + i)
                        if value is not None:
                            ws[f'{col_letter}{current_row}'] = value
                            ws[f'{col_letter}{current_row}'].number_format = '#,##0.00'
                    
                    ws[f'{sum_col}{current_row}'] = data.get('total', 0)
                    ws[f'{sum_col}{current_row}'].number_format = '#,##0.00'
                    ws[f'{sum_col}{current_row}'].font = Font(bold=True)
                    
                    current_row += 1
            
            current_row += 1
            
            # B. Betriebsausgaben
            ws[f'A{current_row}'] = "B. Betriebsausgaben"
            ws[f'A{current_row}'].font = header_font
            current_row += 1
            
            for field, data in self.extracted_data.items():
                if field.startswith('_'):
                    continue
                if field.startswith('B'):
                    ws[f'A{current_row}'] = field
                    ws[f'B{current_row}'] = data['description']
                    
                    for i, value in enumerate(data['values']):
                        col_letter = chr(ord('C') + i)
                        if value is not None:
                            ws[f'{col_letter}{current_row}'] = value
                            ws[f'{col_letter}{current_row}'].number_format = '#,##0.00'
                    
                    ws[f'{sum_col}{current_row}'] = data.get('total', 0)
                    ws[f'{sum_col}{current_row}'].number_format = '#,##0.00'
                    ws[f'{sum_col}{current_row}'].font = Font(bold=True)
                    
                    current_row += 1
            
            # Spaltenbreiten anpassen
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            for i in range(len(months) + 1):
                col_letter = chr(ord('C') + i)
                ws.column_dimensions[col_letter].width = 15
            
            wb.save(export_path)
            return True
            
        except Exception as e:
            print(f"Automatic Export Fehler: {e}")
            return False
    
    def update_customer_history(self):
        if self.current_customer and self.bwa_file_path:
            history_entry = {
                "file_path": os.path.basename(self.bwa_file_path),
                "period": f"{self.selected_start_month}-{self.selected_end_month}",
                "year": self.selected_year,
                "processed_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "confidence": self.calculate_average_confidence()
            }
            
            self.current_customer.bwa_history.append(history_entry)
            self.customer_manager.save_customer(self.current_customer)
    
    def calculate_average_confidence(self) -> float:
        if not self.extracted_data:
            return 0.0
        
        confidences = [data.get('confidence', 0) for field, data in self.extracted_data.items() if not field.startswith('_')]
        return sum(confidences) / len(confidences) if confidences else 0.0
    
    def update_data_value(self, data_row_index: int, month_index: int, new_value: float):
            """EditableLabel'dan gelen geri bildirimi işler ve veriyi günceller."""
            # Hangi EKS alanının güncellendiğini bul
            fields = sorted([k for k in self.extracted_data.keys() if not k.startswith('_')])
            field_key = fields[data_row_index - 1] # -1 çünkü başlık satırı var
            
            # Arka plan verisini güncelle
            self.extracted_data[field_key]['values'][month_index] = new_value
            
            # Satır toplamını yeniden hesapla
            new_total = sum(v for v in self.extracted_data[field_key]['values'] if v is not None)
            self.extracted_data[field_key]['total'] = new_total
            
            # Arayüzdeki toplam etiketini widget referansı üzerinden güncelle
            if field_key in self.total_labels:
                self.total_labels[field_key].update_text(new_total)
    

class CustomerDialog(ctk.CTkToplevel):
    def __init__(self, parent, texts):
        super().__init__(parent)
        
        self.texts = texts
        self.result = None
        
        self.title("Neuer Kunde")
        self.geometry("400x300")
        self.configure(fg_color="#2b2b2b")
        
        self.transient(parent)
        self.grab_set()
        
        self.setup_ui()
        self.center_window()
    
    def setup_ui(self):
        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        title_label = ctk.CTkLabel(main_frame, text="Neuen Kunden erstellen", 
                                 font=ctk.CTkFont(size=18, weight="bold"))
        title_label.pack(pady=20)
        
        code_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        code_frame.pack(fill="x", pady=10)
        
        ctk.CTkLabel(code_frame, text="Kundennummer:", width=120).pack(side="left")
        self.code_entry = ctk.CTkEntry(code_frame, width=200)
        self.code_entry.pack(side="right")
        
        name_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        name_frame.pack(fill="x", pady=10)
        
        ctk.CTkLabel(name_frame, text="Kundenname:", width=120).pack(side="left")
        self.name_entry = ctk.CTkEntry(name_frame, width=200)
        self.name_entry.pack(side="right")
        
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(side="bottom", pady=20)
        
        cancel_btn = ctk.CTkButton(button_frame, text="Abbrechen", 
                                 command=self.cancel, width=100)
        cancel_btn.pack(side="left", padx=10)
        
        save_btn = ctk.CTkButton(button_frame, text="Speichern", 
                               command=self.save, width=100)
        save_btn.pack(side="right", padx=10)
        
        self.code_entry.focus()
    
    def center_window(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (400 // 2)
        y = (self.winfo_screenheight() // 2) - (300 // 2)
        self.geometry(f"400x300+{x}+{y}")
    
    def save(self):
        code = self.code_entry.get().strip()
        name = self.name_entry.get().strip()
        
        if not code or not name:
            messagebox.showwarning("Warnung", "Bitte alle Felder ausfüllen")
            return
        
        self.result = {
            "code": code,
            "name": name
        }
        self.destroy()
    
    def cancel(self):
        self.destroy()


class SettingsDialog(ctk.CTkToplevel):
    def __init__(self, parent, texts):
        super().__init__(parent)
        
        self.texts = texts
        self.settings_file = "settings.json"
        self.settings = self.load_settings()
        
        self.title("Einstellungen")
        self.geometry("500x400")
        self.configure(fg_color="#2b2b2b")
        
        self.transient(parent)
        self.grab_set()
        
        self.setup_ui()
        self.center_window()
    
    def load_settings(self) -> Dict:
        try:
            if os.path.exists(self.settings_file):
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception:
            pass
        
        return {
            "claude_api_key": "",
            "auto_customer_creation": True,
            "default_template": "eks_standard.xlsx",
            "backup_enabled": True
        }
    
    def save_settings(self):
        try:
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(self.settings, f, ensure_ascii=False, indent=2)
            return True
        except Exception:
            return False
    
    def setup_ui(self):
        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        title_label = ctk.CTkLabel(main_frame, text="Einstellungen", 
                                 font=ctk.CTkFont(size=18, weight="bold"))
        title_label.pack(pady=20)
        
        settings_frame = ctk.CTkScrollableFrame(main_frame, fg_color="#3b3b3b")
        settings_frame.pack(fill="both", expand=True, pady=10)
        
        # API Einstellungen
        api_section = ctk.CTkFrame(settings_frame, fg_color="#4b4b4b")
        api_section.pack(fill="x", pady=10, padx=10)
        
        ctk.CTkLabel(api_section, text="API Einstellungen", 
                    font=ctk.CTkFont(size=14, weight="bold")).pack(pady=10)
        
        api_frame = ctk.CTkFrame(api_section, fg_color="transparent")
        api_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(api_frame, text="Claude API Key:", width=150).pack(side="left")
        self.api_key_entry = ctk.CTkEntry(api_frame, width=250, show="*")
        self.api_key_entry.pack(side="right", padx=10)
        self.api_key_entry.insert(0, self.settings.get("claude_api_key", ""))
        
        test_btn = ctk.CTkButton(api_section, text="API Testen", command=self.test_api, width=100)
        test_btn.pack(pady=10)
        
        # Allgemeine Einstellungen
        general_section = ctk.CTkFrame(settings_frame, fg_color="#4b4b4b")
        general_section.pack(fill="x", pady=10, padx=10)
        
        ctk.CTkLabel(general_section, text="Allgemeine Einstellungen", 
                    font=ctk.CTkFont(size=14, weight="bold")).pack(pady=10)
        
        self.auto_customer_var = ctk.BooleanVar(value=self.settings.get("auto_customer_creation", True))
        auto_customer_check = ctk.CTkCheckBox(general_section, text="Kunden automatisch aus BWA erstellen",
                                            variable=self.auto_customer_var)
        auto_customer_check.pack(pady=5, padx=20, anchor="w")
        
        self.backup_var = ctk.BooleanVar(value=self.settings.get("backup_enabled", True))
        backup_check = ctk.CTkCheckBox(general_section, text="Automatische Backups aktiviert",
                                     variable=self.backup_var)
        backup_check.pack(pady=5, padx=20, anchor="w")
        
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(side="bottom", pady=20)
        
        cancel_btn = ctk.CTkButton(button_frame, text="Abbrechen", 
                                 command=self.cancel, width=100)
        cancel_btn.pack(side="left", padx=10)
        
        save_btn = ctk.CTkButton(button_frame, text="Speichern", 
                               command=self.save, width=100)
        save_btn.pack(side="right", padx=10)
    
    def center_window(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (500 // 2)
        y = (self.winfo_screenheight() // 2) - (400 // 2)
        self.geometry(f"500x400+{x}+{y}")
    
    def test_api(self):
        """API test fonksiyonu - DÜZELTİLMİŞ"""
        api_key = self.api_key_entry.get().strip()
        
        if not api_key:
            messagebox.showwarning("Warnung", "Bitte API Key eingeben")
            return
        
        if not api_key.startswith('sk-ant-'):
            messagebox.showwarning("Warnung", 
                "API Key Format scheint falsch zu sein.\n"
                "Claude API Keys beginnen mit 'sk-ant-api03-'")
            return
        
        progress = ctk.CTkToplevel(self)
        progress.title("API Test")
        progress.geometry("400x150")
        progress.transient(self)
        progress.grab_set()
        
        progress.update_idletasks()
        x = (progress.winfo_screenwidth() // 2) - 200
        y = (progress.winfo_screenheight() // 2) - 75
        progress.geometry(f"400x150+{x}+{y}")
        
        status_label = ctk.CTkLabel(progress, text="API wird getestet...", 
                                    font=ctk.CTkFont(size=14))
        status_label.pack(pady=20)
        
        detail_label = ctk.CTkLabel(progress, text="", font=ctk.CTkFont(size=11))
        detail_label.pack(pady=5)
        
        progress_bar = ctk.CTkProgressBar(progress)
        progress_bar.pack(pady=10, padx=20, fill="x")
        progress_bar.start()
        
        def test_thread():
            try:
                detail_label.configure(text="Verbindung zu Claude API...")
                
                headers = {
                    "Content-Type": "application/json",
                    "x-api-key": api_key,  # x-api-key kullan
                    "anthropic-version": "2023-06-01"
                }
                
                data = {
                    "model": "claude-3-haiku-20240307",
                    "max_tokens": 10,
                    "messages": [
                        {
                            "role": "user",
                            "content": "Say 'OK'"
                        }
                    ]
                }
                
                response = requests.post(
                    "https://api.anthropic.com/v1/messages",
                    headers=headers,
                    json=data,
                    timeout=10
                )
                
                self.after(0, lambda: progress_bar.stop())
                
                if response.status_code == 200:
                    self.after(0, lambda: detail_label.configure(text="✅ Erfolgreich!"))
                    import time
                    time.sleep(1)
                    self.after(0, lambda: progress.destroy())
                    self.after(0, lambda: messagebox.showinfo("Erfolg", 
                        "✅ Claude API Verbindung erfolgreich!\n"
                        "Sie können jetzt AI-Vorschläge nutzen."))
                elif response.status_code == 401:
                    self.after(0, lambda: progress.destroy())
                    self.after(0, lambda: messagebox.showerror("Fehler", 
                        "❌ Ungültiger API Key!\n"
                        "Bitte überprüfen Sie Ihren API Key."))
                elif response.status_code == 429:
                    self.after(0, lambda: progress.destroy())
                    self.after(0, lambda: messagebox.showerror("Fehler", 
                        "⚠️ Rate Limit erreicht!\n"
                        "Bitte warten Sie einen Moment."))
                else:
                    error_detail = response.json().get('error', {}).get('message', f'Status: {response.status_code}')
                    self.after(0, lambda: progress.destroy())
                    self.after(0, lambda: messagebox.showerror("Fehler", 
                        f"❌ API Test fehlgeschlagen:\n{error_detail}"))
                    
            except requests.exceptions.Timeout:
                self.after(0, lambda: progress.destroy())
                self.after(0, lambda: messagebox.showerror("Fehler", 
                    "⏱️ Zeitüberschreitung!\n"
                    "Bitte versuchen Sie es später erneut."))
            except Exception as e:
                self.after(0, lambda: progress.destroy())
                self.after(0, lambda: messagebox.showerror("Fehler", 
                    f"❌ Verbindungsfehler:\n{str(e)}"))
        
        threading.Thread(target=test_thread, daemon=True).start()
    
    def save(self):
        """Settings kaydetme - DÜZELTİLMİŞ"""
        api_key = self.api_key_entry.get().strip()
        
        if api_key and not api_key.startswith('sk-ant-'):
            messagebox.showwarning("Warnung", "API Key Format ungültig. Sollte mit 'sk-ant-' beginnen")
            return
        
        self.settings["claude_api_key"] = api_key
        self.settings["auto_customer_creation"] = self.auto_customer_var.get()
        self.settings["backup_enabled"] = self.backup_var.get()
        
        if self.save_settings():
            # Ana penceredeki API'yi güncelle
            if hasattr(self.master, 'bwa_parser'):
                self.master.bwa_parser.set_claude_api(api_key)
            messagebox.showinfo("Erfolg", "Einstellungen gespeichert")
            self.destroy()
        else:
            messagebox.showerror("Fehler", "Einstellungen konnten nicht gespeichert werden")
    
    def cancel(self):
        self.destroy()

class TemplateManager:
    def __init__(self, template_dir: str = "templates"):
        self.template_dir = template_dir
        os.makedirs(template_dir, exist_ok=True)
        self.create_default_template()
    
    def create_default_template(self):
        """Erstellt ein Standard EKS Template falls nicht vorhanden"""
        template_path = os.path.join(self.template_dir, "eks_standard.xlsx")
        if not os.path.exists(template_path):
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "EKS Standard"
                
                ws['A1'] = "EKS Standard Template"
                ws['A2'] = "Dieses Template wird automatisch befüllt"
                
                ws['A4'] = "A. Betriebseinnahmen"
                ws['A5'] = "A1 - Betriebseinnahmen"
                ws['A6'] = "A5 - Umsatzsteuer"
                
                ws['A8'] = "B. Betriebsausgaben"
                ws['A9'] = "B1 - Material, Stoffe, Waren"
                ws['A10'] = "B2c - Aushilfslöhne"
                ws['A11'] = "B3 - Miete und Energiekosten"
                ws['A12'] = "B11 - Telefon"
                ws['A13'] = "B14c - Nebenkosten Geldverkehr"
                ws['A14'] = "B17 - Vorsteuer"
                
                wb.save(template_path)
            except Exception as e:
                print(f"Fehler beim Erstellen des Standard Templates: {e}")
    
    def get_available_templates(self) -> List[str]:
        """Gibt Liste verfügbarer Templates zurück"""
        templates = []
        for file in os.listdir(self.template_dir):
            if file.endswith('.xlsx'):
                templates.append(file)
        return sorted(templates)

class EditableLabel:
    """Tıklandığında CTkEntry'ye dönüşen, fare etkileşimli bir etiket widget'ı."""
    def __init__(self, master, row, column, text, callback, is_total=False):
        self.master = master
        self.row = row
        self.column = column
        self.callback = callback
        
        # Ana widget'ın arka plan rengini al
        self.original_bg = master.cget("fg_color")
        
        self.value = text if isinstance(text, (int, float)) else 0.0
        display_text = f"{self.value:,.2f} €" if text is not None else "N/A"
        
        font_weight = "bold" if is_total else "normal"
        self.font = ctk.CTkFont(size=12, weight=font_weight)
        
        # Etiketi bir çerçeve içine yerleştirerek daha iyi kontrol sağlıyoruz
        self.frame = ctk.CTkFrame(master, fg_color="transparent")
        self.frame.grid(row=row, column=column, sticky="ew")
        
        self.label = ctk.CTkLabel(self.frame, text=display_text, font=self.font, anchor="e")
        self.label.pack(fill="x", padx=10, pady=8) # Dikey ve yatay boşluk ekle
        
        # Etkileşim için olayları bağla
        self.label.bind("<Button-1>", self._on_click)
        self.label.bind("<Enter>", self._on_enter)
        self.label.bind("<Leave>", self._on_leave)
        
        self.entry = None

    def _on_enter(self, event):
        """Fare üzerine geldiğinde görsel geri bildirim verir."""
        self.label.configure(fg_color="#3a3a3a", cursor="hand2")

    def _on_leave(self, event):
        """Fare ayrıldığında eski haline döner."""
        self.label.configure(fg_color="transparent", cursor="")

    def _on_click(self, event):
        """Etikete tıklandığında düzenleme modunu başlatır."""
        # Mevcut değeri al
        clean_text = f"{self.value:.2f}"
        
        self.entry = ctk.CTkEntry(self.frame, font=self.font, justify="right")
        self.entry.insert(0, clean_text)
        self.entry.pack(fill="x", padx=10, pady=8)
        self.entry.focus_set()
        
        self.label.pack_forget() # Etiketi gizle
        
        self.entry.bind("<Return>", self._on_save)
        self.entry.bind("<FocusOut>", self._on_save)

    def _on_save(self, event):
        """Değeri kaydeder ve tekrar etiket moduna döner."""
        new_value_str = self.entry.get().replace(",", ".")
        try:
            self.value = float(new_value_str)
            # Ana uygulamadaki veriyi güncellemek için callback fonksiyonunu çağır
            self.callback(self.row, self.column - 2, self.value)
        except (ValueError, TypeError):
            pass # Geçersiz giriş varsa, değişikliği yoksay

        self.label.configure(text=f"{self.value:,.2f} €")
        self.entry.destroy()
        self.label.pack(fill="x", padx=10, pady=8)

    def update_text(self, new_value):
        """Dışarıdan değeri güncellemek için kullanılır."""
        self.value = new_value
        self.label.configure(text=f"{self.value:,.2f} €")


# Hauptprogramm
def main():
    # Arbeitsverzeichnisse erstellen
    directories = ["data", "data/customers", "templates", "exports"]
    for directory in directories:
        os.makedirs(directory, exist_ok=True)
    
    # CustomTkinter Erscheinungsbild setzen
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    
    # Hauptanwendung starten
    app = EKSFormFiller()
    app.mainloop()


if __name__ == "__main__":
    main()